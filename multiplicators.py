import time
import requests
from bs4 import BeautifulSoup
import random
import os
from openpyxl import load_workbook
import re

def get_proxy():
    proxy_url =  "https://free-proxy-list.net"
    page = requests.get(proxy_url)
    html = BeautifulSoup(page.content, "html.parser")
    table = html.find(id="proxylisttable")
    # print("Table: %s" % table.prettify())
    http_list = []
    https_list = []
    for row in table.tbody.findAll("tr"):
        ip = row.findAll("td")[0].contents[0]
        port = row.findAll("td")[1].contents[0]
        https = row.findAll("td")[6].contents[0]
        if https == "yes":
            https_list.append("%s:%s" % (ip, port))
        else:
            http_list.append("%s:%s" % (ip, port))
    return http_list, https_list

def clean_value(value):
    m = re.search('\d+[.|,]?\d+', value)
    if m:
        value = m.group(0)
    value = value.replace(" ", "").replace("\n", "").replace("\\n", "").replace("\\", "")
    return value

def get_summary_table_value(html, multiplicator):
    result = 0
    try:
        stock_summary_table = html.find("div", class_="stock-summary-table fc-regular")
        summary_values = stock_summary_table.findAll("div")
        summary_values_len = len(summary_values)
        i = 0
        for _ in stock_summary_table.findAll("div"):
            i += 1
            value = _.contents[0]
            if multiplicator in value:
                i = summary_values_len - 1
            if i == summary_values_len:
                result = clean_value(value)
                break
    except Exception as err:
        pass
    return result

def get_indicator_table_value(html, table_type, multiplicator, exclude=None):
    value = 0
    try:
        div = html.find("div", id=table_type)
        table = div.find("table", class_="stock-indicator-table")
        for row in table.tbody.findAll("tr"):
            a = row.findAll("td")[0].contents[0]
            name = a.contents[0]
            if exclude:
                if exclude in name:
                    continue
            if multiplicator in name:
                value = clean_value(row.findAll("td")[1].contents[0])
    except Exception as err:
        pass
    return value

def get_stock_statistics_value(html, multiplicator):
    value = 0
    try:
        statistic = html.find("div", id="stock-statistics")
        div = statistic.find("div")
        for row in div.findAll("div", class_="statictics-item"):
            name = row.contents[0]
            if multiplicator in name:
                value = clean_value(row.find("span").contents[0])
    except Exception as err:
        pass
    return value

def parse_gurufocus_page(url, proxy=True, tries=10):
    if proxy:
        http_list, https_list = get_proxy()
        proxy_list = https_list if url.startswith("https") else http_list
        if not proxy_list:
            raise Exception("Empty proxy list")
        random.shuffle(proxy_list)
        key = "https" if url.startswith("https") else "http"
        for proxy in proxy_list:
            try:
                r = requests.get(url=url, proxies={key: "http://%s" % proxy}, timeout=3)
                break
            except Exception as err:
                # print(err)
                print("%s proxy %s FAILED. Continue..." % (key, proxy))
                time.sleep(2)
                continue
    else:
        r = requests.get(url=url)

    content = str(r.content)
    necessary_attributes = [
        "class=\"stock-indicator-table\"",
        "class=\"stock-summary-table fc-regular\""
        "id=\"stock-statistics\"",
    ]

    # OPEN FILE:
    # with open("GuruFocus.com.html", "r", errors="ignore") as _f:
    #     content = str(_f.read())

    # SELENIUM:
    # from selenium import webdriver
    # op = webdriver.ChromeOptions()
    # op.add_argument('headless')
    # driver = webdriver.Chrome(options=op)
    # driver.get(url)
    # time.sleep(3)
    # content = str(driver.page_source)

    for attribute in necessary_attributes:
        if attribute not in content and tries > 0:
           #RECURSION HERE:
           tries -= 1
           print("'%s' atrribute missed. Recursion..." % attribute)
           return parse_gurufocus_page(url, proxy, tries=tries)

    html = BeautifulSoup(content, "html.parser")
    P_E = get_summary_table_value(html, "P/E")
    P_B = get_summary_table_value(html, "P/B")
    MARGIN = get_indicator_table_value(html, "profitability", "Net Margin")
    ROE = get_indicator_table_value(html, "profitability", "ROE %")
    D_E = get_indicator_table_value(html, "financial-strength", "Debt-to-Equity")
    D_EBITDA = get_indicator_table_value(html, "financial-strength", "Debt-to-EBITDA")
    FORWARD_P_E = get_indicator_table_value(html, "ratios", "Forward PE Ratio")
    EPS = get_stock_statistics_value(html, "EPS")
    E_V_EBITDA = get_indicator_table_value(html, "ratios", "EV-to-EBITDA")
    P_S = get_indicator_table_value(html, "ratios", "PS Ratio")
    PEG = get_indicator_table_value(html, "ratios", "PEG Ratio")
    CURRENT_RATIO = get_indicator_table_value(html, "ratios", "Current Ratio")
    QUICK_RATIO = get_indicator_table_value(html, "ratios", "Quick Ratio")
    DIV = get_indicator_table_value(html, "dividend", "Dividend Yield",
                                    exclude="Forward Dividend Yield")
    DIV_PAYOUT_RATIO = get_indicator_table_value(html, "dividend", "Dividend Payout Ratio")
    return locals()


# url = "https://api.ipify.org?format=json" # get_my IP
# url = "http://checkip.dyndns.org" # get_my IP
# url = "http://ipinfo.io" # get_my IP

d = {
    8: "ABBV",
    11: "SNY",
    14: "PFE"
}


for company_position, ticker in d.items():
    url = "https://www.gurufocus.com/stock/%s/summary" % ticker
    current_dir = os.path.dirname(os.path.realpath(__file__))
    file_path = "%s\stocks.xlsx" % current_dir
    result = parse_gurufocus_page(url)
    result = {key: result[key] for key in result.keys() if key.isupper()}
    multipicators = ["P_E", "MARGIN", "FORWARD_P_E", "E_V_EBITDA", "P_B", "P_S", "D_E", "D_EBITDA",
               "PEG", "DIV", "DIV_PAYOUT_RATIO", "QUICK_RATIO", "CURRENT_RATIO", "ROE"]
    values = [result[col] for col in multipicators]
    wb = load_workbook(filename=file_path)
    ws = wb["Pharmacy"]
    for index in range(3, 17):
        ws.cell(row=index,column=company_position).value = float(values[index-3])
    wb.save(file_path)
    print("%s: - Done" % ticker)
